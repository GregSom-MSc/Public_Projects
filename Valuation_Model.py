# VALUATION MODEL

# This python file produces a Simple Valuation Model Template.

# STEPSL

# Import the necessary stuff
import pandas as pd
import openpyxl as xl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter, quote_sheetname, absolute_coordinate
from openpyxl.utils import cell
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import ColorScaleRule
# Then, load the function


def generate_valuation_excel(filename="Valuation_Model.xlsx"):
    ''' 
    This function creates a complete Excel-based valuation model with three integrated sheets:

    1. Assumptions - Contains configurable model parameters
    2. Valuation Model - Generated calculations and projections over 5-year period
    3. Dashboard - Simple summary of key metrics and performance indicators

    The model includes transaction assumptions, sources & uses of funds, financial projections,
    debt schedules, and exit analysis with IRR calculations.

    Parameters:
    -----------
    filename : str, optional
        The name of the Excel file to create. Default: "Valuation_Model.xlsx".
        IMPORTANT: Remember to include the .xlsx extension.

    Returns:
    --------
    None
        The function saves the Excel file to disk with the specified filename
        and prints a message confirming it.

    Examples:
    ---------
    >>> generate_valuation_excel()  # Creates "Valuation_Model.xlsx"
    >>> generate_valuation_excel("PE_Deal_2025_SGMS.xlsx")  # Creates custom named file

    Notes:
    ------
    - Default assumptions can be modified in the corresponding sheet
    - Financial projections span 5 years, you can configure growth rates
    - IRR calculation uses XIRR function for more accurate returns analysis
    - The model follows PE/M&A industry standard practices

    SGMS 2025
    '''
    # First, we create a new workbook in memory
    wb = xl.Workbook()

    # Let's create sheets: Assumptions
    assumptions_sheet = wb.active
    assumptions_sheet.title = "Assumptions"

    # Let's create sheets: Model
    valuation_sheet = wb.create_sheet(title="Valuation Model")

    # ----- ASSUMPTIONS SHEET ------------------------------------------

    # Adding headers
    assumptions_sheet['A1'] = "Valuation Model Assumptions"
    assumptions_sheet['A1'].font = Font(bold=True, size=14)
    assumptions_sheet['A1'].fill = PatternFill(fgColor="134074")
    assumptions_sheet.merge_cells('A1:C1')
    assumptions_sheet['A1'].alignment = Alignment(horizontal='center')

    # Now, we define the sheet's categories and values
    assumptions_data = [
        ["Transaction Assumptions", "", ""],
        ["Purchase Price ($M)", "purchase_price", 100],
        ["EBITDA Multiple", "ebitda_multiple", 8.0],
        ["Revenue Multiple", "revenue_multiple", 2.0],
        ["", "", ""],
        ["Financing Assumptions", "", ""],
        ["Debt Financing (%)", "debt_financing", 0.6],
        ["Equity Financing (%)", "equity_financing", 0.4],
        ["Synergies ($M)", "synergies", 5],
        ["Interest Rate (%)", "interest_rate", 0.05],
        ["Principal Repayment Rate (%)", "principal_rate", 0.1],
        ["Transaction Fees ($M)", "transaction_fees", 2],
        ["Debt to be Refinanced ($M)", "refinanced_debt", 0],
        ["", "", ""],
        ["Operating Assumptions", "", ""],
        ["EBITDA Margin (%)", "ebitda_margin", 0.2],
        ["Tax Rate (%)", "tax_rate", 0.25],
        ["CapEx (% of Revenue)", "capex_rate", 0.05],
        ["Working Capital (% of Revenue)", "wc_rate", 0.1],
        ["Annual Revenue Growth (%)", "revenue_growth", 0.05],
        ["Initial Revenue ($M)", "initial_revenue", 100],
        ["", "", ""],
        ["Exit Assumptions", "", ""],
        ["Exit Year", "exit_year", 5],
        ["Exit EBITDA Multiple", "exit_multiple", 8.0],
    ]

    # Add assumption data
    for row_idx, row_data in enumerate(assumptions_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            # Convert row and column indices to a cell coordinate string
            cell_coord = get_column_letter(col_idx) + str(row_idx)

            # Check if the cell is within a merged cell range
            is_merged = False
            for merged_range in assumptions_sheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = range_boundaries(
                    str(merged_range))
                col_num = col_idx
                row_num = row_idx
                if min_col <= col_num <= max_col and min_row <= row_num <= max_row:
                    is_merged = True
                    break

            if not is_merged:
                cell_obj = assumptions_sheet.cell(
                    row=row_idx, column=col_idx, value=value)
            else:
                merged_cell_range = None
                for merged_range in assumptions_sheet.merged_cells.ranges:
                    min_col, min_row, max_col, max_row = range_boundaries(
                        str(merged_range))
                    col_num = col_idx
                    row_num = row_idx
                    if min_col <= col_num <= max_col and min_row <= row_num <= max_row:
                        merged_cell_range = merged_range
                        break
                if (row_idx, col_idx) == (merged_cell_range.min_row, merged_cell_range.min_col):
                    cell_obj = assumptions_sheet.cell(
                        row=row_idx, column=col_idx, value=value)
                else:
                    continue

            # Format headers
            if col_idx == 1 and not value.startswith("  ") and value != "":
                if "Assumptions" in value:
                    cell_obj.font = Font(bold=True, size=12, color="FFFFFF")
                    cell_obj.fill = PatternFill("solid", fgColor="134074")
                    assumptions_sheet.merge_cells(f'A{row_idx}:C{row_idx}')
                    cell_obj.alignment = Alignment(horizontal='left')
                else:
                    cell_obj.font = Font(bold=True)

            # Format percentages and numbers
            if col_idx == 3 and isinstance(value, (int, float)):
                if "(%)" in row_data[0]:
                    cell_obj.number_format = '0.0%'
                else:
                    cell_obj.number_format = '#,##0.0'

    # Let's define named ranges for easier formula references
    for row in assumptions_data:
        if len(row) >= 3 and isinstance(row[1], str) and row[1] != "":
            name = row[1]
            row_index = assumptions_data.index(row) + 2
            ref = f'{quote_sheetname(assumptions_sheet.title)}!{absolute_coordinate(f"C{row_index}")}'
            defn = DefinedName(name, attr_text=ref)
            wb.defined_names[name] = defn

    # Format the assumptions sheet
    assumptions_sheet.column_dimensions['A'].width = 30
    assumptions_sheet.column_dimensions['B'].width = 25
    assumptions_sheet.column_dimensions['C'].width = 15

# -------------------------------------------------------------------------------------
######################################################################################

    # ----- VALUATION MODEL SHEET -----
    # Set up Valuation Model Structure
    sections = [
        {
            "title": "Transaction Assumptions",
            "rows": [
                {"concept": "Purchase Price ($M)",
                 "formula": "=purchase_price"},
                {"concept": "EBITDA Multiple", "formula": "=ebitda_multiple"},
                {"concept": "Revenue Multiple", "formula": "=revenue_multiple"},
                {"concept": "", "formula": ""},
                {"concept": "Debt Financing (%)",
                 "formula": "=debt_financing"},
                {"concept": "Equity Financing (%)",
                 "formula": "=equity_financing"},
                {"concept": "Synergies ($M)", "formula": "=synergies"}
            ]
        },
        {
            "title": "Sources & Uses",
            "rows": [
                {"concept": "Debt Raised ($M)",
                 "formula": "=purchase_price*debt_financing"},
                {"concept": "Equity Raised ($M)",
                 "formula": "=purchase_price*equity_financing"},
                {"concept": "Total Sources ($M)", "formula": "=D12+D13"},
                {"concept": "", "formula": ""},
                {"concept": "Purchase Price ($M)",
                 "formula": "=purchase_price"},
                {"concept": "Transaction Fees ($M)",
                 "formula": "=transaction_fees"},
                {"concept": "Debt Refinanced ($M)",
                 "formula": "=refinanced_debt"},
                {"concept": "Total Uses ($M)", "formula": "=SUM(D16:D18)"}
            ]
        },
        {
            "title": "Financial Projections",
            "rows": [
                {"concept": "Revenue ($M)", "formula": "=initial_revenue"},
                {"concept": "EBITDA ($M)", "formula": "=E23*ebitda_margin"},
                {"concept": "Net Income ($M)", "formula": "=E24*(1-tax_rate)"},
                {"concept": "CapEx ($M)", "formula": "=E23*capex_rate"},
                {"concept": "Working Capital ($M)", "formula": "=E23*wc_rate"}
            ]
        },
        {
            "title": "Debt Schedule",
            "rows": [
                {"concept": "Beginning Debt ($M)", "formula": "=D13"},
                {"concept": "Interest Expense ($M)",
                 "formula": "=E30*interest_rate"},
                {"concept": "Principal Payment ($M)",
                 "formula": "=E30*principal_rate"},
                {"concept": "Ending Debt ($M)", "formula": "=E30-E32"}
            ]
        },
        {
            "title": "Exit Analysis",
            "rows": [
                {"concept": "Exit Year", "formula": "=exit_year"},
                {"concept": "Exit EBITDA Multiple", "formula": "=exit_multiple"},
                {"concept": "Exit Enterprise Value ($M)",
                 "formula": "=D36*I23"},
                {"concept": "Debt Remaining ($M)", "formula": "=I33"},
                {"concept": "Equity Value ($M)", "formula": "=D37-D38"},
                {"concept": "IRR (%)",
                 "formula": "=XIRR(D14*-1,D40,TODAY(),TODAY()+365*exit_year)"}
            ]
        }
    ]

    # Add headers to valuation sheet
    valuation_sheet['A1'] = "Concept"
    valuation_sheet['B1'] = "Description"
    valuation_sheet['A1'].font = Font(bold=True)
    valuation_sheet['B1'].font = Font(bold=True)
    valuation_sheet['A1'].fill = PatternFill(fgColor="134074")
    valuation_sheet['B1'].fill = PatternFill(fgColor="134074")

    # Add year headers
    for year in range(1, 6):
        col = get_column_letter(3 + year)
        valuation_sheet[f"{col}1"] = f"Year {year}"
        valuation_sheet[f"{col}1"].font = Font(bold=True)
        valuation_sheet[f"{col}1"].fill = PatternFill(fgColor="134074")
        valuation_sheet[f"{col}1"].alignment = Alignment(horizontal='center')

    # Current row tracker
    current_row = 2

    # Populate the valuation model sheet with sections
    for section in sections:
        # Add section title
        valuation_sheet[f'A{current_row}'] = section["title"]
        valuation_sheet[f'A{current_row}'].font = Font(bold=True)
        valuation_sheet[f'A{current_row}'].fill = PatternFill(
            "solid", fgColor="D9D9D9")
        valuation_sheet.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1

        # Add rows for this section
        for row_data in section["rows"]:
            valuation_sheet[f'A{current_row}'] = row_data["concept"]

            # Add formula for initial column (usually year 0 or base value)
            if row_data["formula"]:
                valuation_sheet[f'D{current_row}'] = row_data["formula"]

            current_row += 1

        # Add blank row after section
        current_row += 1

    # Add year projections for revenue
    for year in range(2, 6):
        col = get_column_letter(3 + year)
        prev_col = get_column_letter(2 + year)
        valuation_sheet[f"{col}22"] = f"={prev_col}22*(1+revenue_growth)"

    # Add projections for other financial metrics
    for year in range(2, 6):
        col = get_column_letter(4 + year)
        # EBITDA
        valuation_sheet[f"{col}23"] = f"={col}22*ebitda_margin"
        # Net Income
        valuation_sheet[f"{col}24"] = f"={col}23*(1-tax_rate)"
        # CapEx
        valuation_sheet[f"{col}25"] = f"={col}22*capex_rate"
        # Working Capital
        valuation_sheet[f"{col}26"] = f"={col}22*wc_rate"

    # Add debt schedule projections
    for year in range(2, 6):
        col = get_column_letter(3 + year)
        prev_col = get_column_letter(2 + year)

        # Beginning Debt
        valuation_sheet[f"{col}29"] = f"={prev_col}32"
        # Interest Expense
        valuation_sheet[f"{col}30"] = f"={col}29*interest_rate"
        # Principal Payment
        valuation_sheet[f"{col}31"] = f"={col}29*principal_rate"
        # Ending Debt
        valuation_sheet[f"{col}32"] = f"={col}29-{col}31"

    # Format the valuation sheet
    # Adjust column widths
    valuation_sheet.column_dimensions['A'].width = 25
    valuation_sheet.column_dimensions['B'].width = 25

    for col in range(4, 10):
        valuation_sheet.column_dimensions[get_column_letter(col)].width = 15

    # Add cell formatting for percentages and numbers
    for row in range(1, 50):
        for col in range(4, 10):
            col_letter = get_column_letter(col)
            cell = valuation_sheet[f"{col_letter}{row}"]
            if cell.value:
                concept_cell = valuation_sheet[f"A{row}"]
                if concept_cell.value and "(%)" in str(concept_cell.value):
                    cell.number_format = '0.0%'
                elif isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    cell.number_format = '#,##0.0'

    # Add conditional formatting for financial projections
    projection_rows = [22, 23, 24]  # Revenue, EBITDA, Net Income
    for row in projection_rows:
        valuation_sheet.conditional_formatting.add(
            f'E{row}:I{row}',
            ColorScaleRule(start_type='min', start_color='FFFFFF',
                           end_type='max', end_color='44FF44')
        )

# -------------------------------------------------------------------------------------
######################################################################################

# ---- DASHBOARD SHEET -----
    # Let's add Dashboard sheet
    dashboard_sheet = wb.create_sheet(title="Dashboard")

    # Add title
    dashboard_sheet['A1'] = "Valuation Model Dashboard"
    dashboard_sheet['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    dashboard_sheet['A1'].fill = PatternFill(fgColor="134074")
    dashboard_sheet.merge_cells('A1:F1')
    dashboard_sheet['A1'].alignment = Alignment(horizontal='center')

    # Add key metrics
    dashboard_metrics = [
        ["Purchase Price ($M)", "=purchase_price"],
        ["Debt / Equity Ratio", "=debt_financing/equity_financing"],
        ["Initial EBITDA Margin", "=ebitda_margin"],
        ["5-Year Revenue CAGR", "=(I23/E23)^(1/4)-1"],
        ["Exit Enterprise Value ($M)", "='Valuation Model'!D38"],
        ["Equity Value at Exit ($M)", "='Valuation Model'!D40"],
        ["IRR (%)", "='Valuation Model'!D41"],
        ["Money Multiple", "='Valuation Model'!D40/'Valuation Model'!D14"]
    ]

    # Add metrics to dashboard
    dashboard_sheet['A3'] = "Key Metrics"
    dashboard_sheet['A3'].font = Font(bold=True, size=12)
    dashboard_sheet.merge_cells('A3:B3')

    for idx, (metric, formula) in enumerate(dashboard_metrics, start=4):
        dashboard_sheet[f'A{idx}'] = metric
        dashboard_sheet[f'B{idx}'] = formula

        # Format cells
        if "(%)" in metric:
            dashboard_sheet[f'B{idx}'].number_format = '0.0%'
        else:
            dashboard_sheet[f'B{idx}'].number_format = '#,##0.0'

    # Adjust column widths in dashboard
    dashboard_sheet.column_dimensions['A'].width = 25
    dashboard_sheet.column_dimensions['B'].width = 15

    # Save the workbook
    wb.save(filename)
    print(f"Valuation model generated: {filename}")


#################################################################################################################
# Finally, our script will run the function and automatically create the excel blueprint to work on our Valuation:

if __name__ == "__main__":
    generate_valuation_excel()
